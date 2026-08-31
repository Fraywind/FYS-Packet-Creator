"""Write true end-of-term enrollment into the PDF 7 input spreadsheet.

Page 6 and page 7 both print a column headed "Enrolled", but the two years mean
different things by it, and that difference is the whole reason this script
exists.

  Page 6 reports the UPCOMING year. Those seminars have not run, so the only
  number that exists is the lottery placement, and pdf6_enrollment.py reads
  "Placed" on purpose. The page carries a legend saying so.

  Page 7 reports the PAST year. That year is over and the registrar has real
  counts, so "Enrolled" must mean students actually enrolled, not students
  placed in August.

The 2025-2026 build got this wrong. The fall Q file carries a Students_Enrolled
column, so fall came out right, but the spring Q file carries only Q averages,
so the spring rows fell back to "Placed" from the application system. That put
placement numbers on a page that promises actual enrollment: Homi Bhabha's 63N
printed 15, which was its capacity and its August placement, when 4 students
finished it.

Run this after the registrar publishes the term enrollment export, against the
Q report that was assembled from the Q files, and it replaces the enrollment
column with the registrar's numbers.

    python3 tools/build_pdf7_input.py \
        --qreport ~/Downloads/"Past Year Q Reports 2025-2026.xlsx" \
        --spring ~/Downloads/"Spring '26 Enrollment.xlsx" \
        --out ~/Downloads/"Past Year Q Reports 2025-2026.xlsx"

Which registrar column counts as "enrolled" is a real choice, so it is a flag
rather than a constant. The default is the add/drop column ("Enrollment as of
Feb. 3rd" in the spring export), because that is the roster that sat the course
and produced the Q scores printed beside it. The registration-day column counts
students who registered and then dropped.

Nothing is guessed. Every row it cannot match is listed at the end and left
alone, and a seminar the registrar marks CANCELLED is reported if it is still
present in the Q report rather than silently removed.

It does, though, treat the registrar as the authority on enrolment, so it will
overwrite a number a person put there on purpose. 2025-26 has one: Matthew
Rabin's 73R is 5 in the office's own record against the registrar's 9, and 5 is
what the packet prints. Running this against the spring export changes it back
to 9. Every change is listed before anything is written, and --dry-run writes
nothing, so read the list before saving.
"""

import argparse
import re
import shutil
import sys
from datetime import datetime

import openpyxl

# The registrar export puts a report title on row 1 and the real header on row 2,
# but that has not been stable year to year, so the header row is found by
# looking for the column we need rather than assumed.
HEADER_SEARCH_ROWS = 6

SEM_COL = 'Class Name'
CLASS_COL = 'Class Number'
INSTRUCTOR_COL = 'Instructor'
TITLE_COL = 'Title'
DEFAULT_ENROLL_COL = 'Enrollment as of Feb. 3rd'

# Q report columns, with the spacing the file actually uses.
QR_SEM = 'SEM#'
QR_TERM = 'TERM'
QR_LAST = 'LAST NAME'
QR_ENROLL = 'ENROLL '
QR_CLASS = 'CLASS#'
QR_SEMQ = 'SEMQ'
QR_INSTQ = 'INST Q'

NOT_AVAILABLE = 'N/A'


# A seminar number the source Q file could not supply is written as a placeholder
# rather than left empty, so a placeholder has to read as "absent" or the row
# looks like a seminar number that simply does not match anything.
SEM_PLACEHOLDERS = {'', 'N/A', 'NA', '#N/A', 'NONE', 'TBD', '-', '--'}


def norm_sem(value):
    """FYSEMR 63N, fysemr63n, 63N all reduce to 63N. Placeholders reduce to ''."""
    if value is None:
        return ''
    text = re.sub(r'\s+', ' ', str(value)).strip().upper()
    text = re.sub(r'^FYSEMR\s*', '', text)
    return '' if text in SEM_PLACEHOLDERS else text


def hkey(value):
    """Header names in these files carry stray and doubled spaces ("ENROLL ",
    "FIRST  NAME"). Compare them with the whitespace collapsed so a stray space
    on either side is not a mismatch."""
    return re.sub(r'\s+', ' ', str(value)).strip()


def norm_name(value):
    if value is None:
        return ''
    return re.sub(r'\s+', ' ', str(value)).strip().lower()


def find_header_row(ws, required):
    """Return (row_index, {column name: column index}) for the header row."""
    for row_idx in range(1, min(HEADER_SEARCH_ROWS, ws.max_row) + 1):
        headers = {}
        for col_idx, cell in enumerate(ws[row_idx], start=1):
            if cell.value is not None:
                headers[hkey(cell.value)] = col_idx
        if hkey(required) in headers:
            return row_idx, headers
    raise SystemExit(
        f"Could not find a header row containing {required!r} in the first "
        f"{HEADER_SEARCH_ROWS} rows. Check that the export has not changed shape."
    )


def read_enrollment(path, enroll_col):
    """Read a registrar export into {SEM#: total}, plus the cancelled set.

    A seminar that runs in several sections gets one row per section, so the
    sections are summed: the Q report carries one row per seminar, not per
    section, and page 7 prints the seminar total.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    header_row, headers = find_header_row(ws, SEM_COL)

    for needed in (enroll_col, TITLE_COL):
        if hkey(needed) not in headers:
            raise SystemExit(
                f"{path}: no column named {needed!r}. Found: "
                + ', '.join(sorted(headers))
            )

    totals = {}
    by_class = {}
    sections = {}
    instructors = {}
    cancelled = set()

    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        sem = norm_sem(row[headers[hkey(SEM_COL)] - 1].value)
        if not sem:
            # The trailing totals row has no class name. Skip it.
            continue
        title = str(row[headers[hkey(TITLE_COL)] - 1].value or '')
        if 'CANCELLED' in title.upper():
            cancelled.add(sem)
            continue
        value = row[headers[hkey(enroll_col)] - 1].value
        if value is None:
            continue
        totals[sem] = totals.get(sem, 0) + int(value)
        sections[sem] = sections.get(sem, 0) + 1
        if hkey(CLASS_COL) in headers:
            cls = row[headers[hkey(CLASS_COL)] - 1].value
            if cls not in (None, ''):
                by_class[str(int(float(cls)))] = int(value)
        if hkey(INSTRUCTOR_COL) in headers:
            name = norm_name(row[headers[hkey(INSTRUCTOR_COL)] - 1].value)
            if name:
                instructors.setdefault(name, set()).add(sem)

    return totals, by_class, sections, cancelled, instructors


def apply_term(ws, headers, term, totals, by_class, sections, cancelled,
               instructors, fill_na):
    """Rewrite the enrollment column for one term. Returns a report dict."""
    changed, unchanged, unmatched, recovered, na_filled, still_present = (
        [], [], [], [], [], [])
    multi_section = set()

    for row in ws.iter_rows(min_row=2, values_only=False):
        row_term = str(row[headers[hkey(QR_TERM)] - 1].value or '').strip().lower()
        if row_term != term.lower():
            continue

        sem_cell = row[headers[hkey(QR_SEM)] - 1]
        sem = norm_sem(sem_cell.value)
        last = norm_name(row[headers[hkey(QR_LAST)] - 1].value)

        # A row can reach the Q report without a seminar number when the source
        # Q file had no entry for it. The registrar knows the number, so recover
        # it from the instructor rather than dropping the row.
        if not sem and last:
            candidates = instructors.get(last, set())
            if len(candidates) == 1:
                sem = next(iter(candidates))
                sem_cell.value = sem
                recovered.append((last, sem))

        if not sem:
            unmatched.append((last or '?', 'no seminar number'))
            continue

        if sem in cancelled:
            still_present.append((sem, last))
            continue

        if sem not in totals:
            unmatched.append((sem, last))
            continue

        # A seminar run as several sections shares one SEM#, so the seminar
        # number alone cannot say which section a row belongs to. Where the Q
        # report carries the registrar's CLASS#, use it: that row then gets its
        # own section's enrolment instead of the seminar-wide total.
        cls = row[headers[hkey(QR_CLASS)] - 1].value if hkey(QR_CLASS) in headers else None
        cls = str(int(float(cls))) if cls not in (None, '') else None

        if sections.get(sem, 1) > 1 and cls is None:
            multi_section.add(sem)

        cell = row[headers[hkey(QR_ENROLL)] - 1]
        after = by_class.get(cls) if cls in by_class else totals[sem]
        before = cell.value
        if before == after:
            unchanged.append((sem, last, after))
        else:
            cell.value = after
            changed.append((sem, last, before, after))

        if fill_na:
            for key in (QR_SEMQ, QR_INSTQ):
                q_cell = row[headers[hkey(key)] - 1]
                if q_cell.value in (None, ''):
                    q_cell.value = NOT_AVAILABLE
                    na_filled.append((sem, last, key))

    return {
        'changed': changed, 'unchanged': unchanged, 'unmatched': unmatched,
        'recovered': recovered, 'na_filled': na_filled,
        'still_present': still_present, 'multi_section': sorted(multi_section),
    }


def main():
    parser = argparse.ArgumentParser(
        description='Write true end-of-term enrollment into the PDF 7 input.')
    parser.add_argument('--qreport', required=True,
                        help='the past-year Q report to correct')
    parser.add_argument('--spring', help="registrar's spring enrollment export")
    parser.add_argument('--fall', help="registrar's fall enrollment export")
    parser.add_argument('--enroll-column', default=DEFAULT_ENROLL_COL,
                        help=f'registrar column to treat as enrolled '
                             f'(default: {DEFAULT_ENROLL_COL!r})')
    parser.add_argument('--out', help='where to write (default: in place)')
    parser.add_argument('--no-fill-na', action='store_true',
                        help='leave blank Q scores blank instead of writing N/A')
    parser.add_argument('--dry-run', action='store_true',
                        help='report what would change and write nothing')
    args = parser.parse_args()

    if not args.spring and not args.fall:
        raise SystemExit('Give at least one of --spring or --fall.')

    wb = openpyxl.load_workbook(args.qreport)
    ws = wb.worksheets[0]
    _, headers = find_header_row(ws, QR_SEM)
    for needed in (QR_TERM, QR_ENROLL, QR_LAST, QR_SEMQ, QR_INSTQ):
        if hkey(needed) not in headers:
            raise SystemExit(f'{args.qreport}: no column named {needed!r}.')

    reports = {}
    for term, path in (('Spring', args.spring), ('Fall', args.fall)):
        if not path:
            continue
        totals, by_class, sections, cancelled, instructors = read_enrollment(
            path, args.enroll_column)
        reports[term] = apply_term(ws, headers, term, totals, by_class,
                                   sections, cancelled, instructors,
                                   fill_na=not args.no_fill_na)
        reports[term]['source'] = path
        reports[term]['seminars'] = len(totals)

    for term, rep in reports.items():
        print(f"\n=== {term} ===")
        print(f"source            : {rep['source']}")
        print(f"column used       : {args.enroll_column}")
        print(f"seminars that ran : {rep['seminars']}")
        print(f"enrolment changed : {len(rep['changed'])}")
        print(f"already correct   : {len(rep['unchanged'])}")
        for sem, last, before, after in rep['changed']:
            print(f"    {sem:<5} {last:<14} {before} -> {after}")
        for last, sem in rep['recovered']:
            print(f"  recovered seminar number: {last} -> {sem}")
        for sem, last, key in rep['na_filled']:
            print(f"  no Q score, wrote {NOT_AVAILABLE}: {sem} {last} [{key}]")
        if rep['multi_section']:
            print("  summed across sections: "
                  + ', '.join(rep['multi_section'])
                  + "  (verify the instructors on these)")
        for sem, last in rep['still_present']:
            print(f"  CANCELLED per registrar but present in Q report: "
                  f"{sem} {last}  (review by hand)")
        for sem, last in rep['unmatched']:
            print(f"  NOT MATCHED, left alone: {sem} {last}")

    if args.dry_run:
        print('\nDry run. Nothing written.')
        return

    out = args.out or args.qreport
    if out == args.qreport:
        stamp = datetime.now().strftime('%Y%m%d-%H%M%S')
        backup = args.qreport.replace('.xlsx', f'.backup-{stamp}.xlsx')
        shutil.copy2(args.qreport, backup)
        print(f'\nBacked up to {backup}')
    wb.save(out)
    print(f'Wrote {out}')

    total_unmatched = sum(len(r['unmatched']) for r in reports.values())
    if total_unmatched:
        print(f'\n{total_unmatched} row(s) were not matched and still hold '
              f'their old value. Resolve them before building the packets.')
        sys.exit(1)


if __name__ == '__main__':
    main()
