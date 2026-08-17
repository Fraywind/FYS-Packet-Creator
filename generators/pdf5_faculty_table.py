"""PDF 5: Faculty Teaching Seminars by Name - Table with Checkmarks.

Generates a landscape PDF containing a detailed table listing every faculty member
who has taught a seminar in a specific department, with checkmarks for each academic
year they taught. This is a **department-specific** report.

The table includes:
    - Faculty name (sorted by rank hierarchy, then alphabetically)
    - Faculty rank (with abbreviations like P.O.P. for Professor of the Practice)
    - Checkmark grid for each academic year
    - Color-coded checkmarks to distinguish rank context:
        * Crimson red: Taught under current/most recent rank (standard 'X')
        * Grey: Taught under FYSP entry rank ('X!')
        * Pink: Taught under a higher/promoted rank ('X!!')
    - Special symbols:
        * Double checkmark: Multiple sections taught ('XX')
        * Half circle: Joint department or co-teaching ('X*')

Rank sorting hierarchy (top to bottom):
    1. University Professor
    2. Professor
    3. Associate Professor
    4. Assistant Professor
    5. Professor of the Practice
    6. Lecturer / Senior Lecturer
    7. Emeritus / Emerita
    8. Visiting (all levels)

Data source: SAVED.xlsx (the "Big Faculty Full Spreadsheet")

Dependencies:
    openpyxl: Direct Excel file reading
    reportlab: PDF generation with tables, styled text, and custom symbols
"""

import os
import re
import openpyxl
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch

from .shared import ACRONYM_TO_FULL, trim_years_to_fit, dropped_years


# ---------------------------------------------------------------------------
# Rank Name Conversion
# ---------------------------------------------------------------------------

def get_rank_full_name(rank):
    """Convert a raw rank string to its full display name, handling multi-rank stacking.

    Multi-rank strings (containing '/' or newlines) are split and each part is
    normalized to a standard title. The parts are rejoined with newlines for
    multi-line display in the table cell.

    Args:
        rank: Raw rank string from the spreadsheet. May contain '/' separators
              for rank changes, or be None/empty.

    Returns:
        Normalized rank string with newline separators for multi-rank entries,
        or empty string if rank is None/empty.

    Examples:
        >>> get_rank_full_name('Professor')
        'Professor'
        >>> get_rank_full_name('Assistant Professor/Associate Professor')
        'Assistant Professor\\nAssociate Professor'
        >>> get_rank_full_name(None)
        ''
    """
    if not rank or str(rank).strip() == '':
        return ''

    # Split on '/' or newlines to get individual rank parts
    rank_normalized = str(rank).replace('/', '\n')
    rank_parts = [p.strip() for p in rank_normalized.split('\n') if p.strip()]

    converted = []
    for part in rank_parts:
        pl = part.lower().strip()

        # Match each part to a standard title (most specific first)
        if 'assistant professor' in pl:
            converted.append('Assistant Professor')
        elif 'associate professor' in pl:
            converted.append('Associate Professor')
        elif 'university professor' in pl:
            converted.append('University Professor')
        elif 'professor of the practice' in pl or 'prof of the practice' in pl:
            converted.append('Professor of the Practice')
        elif 'visiting professor' in pl or 'visiting prof' in pl:
            converted.append('Visiting Professor')
        elif 'professor' in pl and not any(x in pl for x in ['associate', 'assistant', 'emeritus', 'visiting', 'practice', 'university']):
            converted.append('Professor')
        elif 'emeritus' in pl or 'emerita' in pl:
            converted.append('Emeritus')
        elif 'senior lecturer' in pl:
            converted.append('Senior Lecturer')
        elif 'lecturer' in pl:
            converted.append('Lecturer')
        elif 'senior preceptor' in pl:
            converted.append('Senior Preceptor')
        elif 'preceptor' in pl:
            converted.append('Preceptor')
        elif 'visiting' in pl:
            converted.append('Visiting')
        elif part.strip():
            converted.append(part.strip())

    return '\n'.join(converted) if converted else str(rank)


# ---------------------------------------------------------------------------
# Rank Sorting
# ---------------------------------------------------------------------------

def get_rank_sort_key(rank):
    """Get a numeric sort key for a faculty rank, used to order rows in the table.

    Returns a numeric value where lower numbers appear first. For multi-rank
    faculty (with '/'), the LAST (most recent) rank determines sort order,
    with special handling for Professor/Emeritus combinations.

    Rank hierarchy (sort order):
        0.5 - University Professor
        1   - Professor
        2   - Associate Professor
        3   - Assistant Professor
        3.5 - Professor of the Practice
        4   - Lecturer / Senior Lecturer
        5   - Emeritus / Emerita (and Professor+Emeritus combos)
        6   - Visiting
        7   - Other / Unknown
        99  - No rank provided

    Args:
        rank: Raw rank string from the spreadsheet. May contain '/' separators.

    Returns:
        Numeric sort key (float). Lower values sort first.
    """
    if not rank:
        return 99

    rank_parts = str(rank).replace('/', '\n').split('\n')
    rank_parts = [p.strip() for p in rank_parts if p.strip()]
    if not rank_parts:
        return 99

    latest = rank_parts[-1].lower()

    # Special case: Professor who also has Emeritus status
    has_prof_emeritus = False
    has_regular_prof = False
    for part in rank_parts:
        pl = part.lower().strip()
        if 'professor' in pl and not any(x in pl for x in ['associate', 'assistant', 'visiting', 'practice', 'university']):
            if any(ep.lower().strip() in ['emeritus', 'emerita'] for ep in rank_parts):
                has_prof_emeritus = True
            else:
                has_regular_prof = True
            break

    # Determine sort key based on most recent rank
    if 'university professor' in latest:
        return 0.5
    elif has_prof_emeritus:
        return 5
    elif has_regular_prof:
        return 1
    elif 'associate professor' in latest and 'visiting' not in latest:
        return 2
    elif 'assistant professor' in latest:
        return 3
    elif 'professor of the practice' in latest or 'prof of the practice' in latest:
        return 3.5
    elif 'lecturer' in latest:
        return 4
    elif 'emeritus' in latest or 'emerita' in latest:
        return 5
    elif 'visiting' in latest:
        return 6
    else:
        return 7


# ---------------------------------------------------------------------------
# Rank Abbreviation
# ---------------------------------------------------------------------------

def get_rank_abbreviation(rank):
    """Abbreviate 'Professor of the Practice' to 'P.O.P.' for compact table display.

    Operates line by line so a multi-rank stack (e.g.
    'Professor of the Practice\\nEmeritus') keeps its other ranks instead of
    collapsing to a single 'P.O.P.'. All other lines are returned unchanged.

    Args:
        rank: Display-ready rank string, possibly newline-separated.

    Returns:
        The rank string with each 'Professor of the Practice' line abbreviated
        to 'P.O.P.' and every other line left as-is.
    """
    if not isinstance(rank, str):
        return str(rank)
    abbreviated_lines = [
        'P.O.P.' if 'professor of the practice' in line.lower() else line
        for line in rank.split('\n')
    ]
    return '\n'.join(abbreviated_lines)


# ---------------------------------------------------------------------------
# Excel Data Reader
# ---------------------------------------------------------------------------

def read_excel_data_by_department(file_path):
    """Read SAVED.xlsx and return data grouped by department.

    This is a duplicate of the same function in pdf4_rank_aggregator.py,
    kept here for module independence. Both produce the same output format.

    Args:
        file_path: Path to the SAVED.xlsx file.

    Returns:
        A tuple of (headers, data_by_dept) where:
        - headers: list of column header strings
        - data_by_dept: dict mapping department acronym -> list of row dicts
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value) if cell.value else f"Column_{cell.column}")

    data_by_dept = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        row_data = {}
        for i, cell in enumerate(row):
            if i < len(headers):
                row_data[headers[i]] = cell.value if cell.value is not None else ""
        dept = row_data.get('Department', '')
        if dept:
            data_by_dept.setdefault(dept, []).append(row_data)

    return headers, data_by_dept


# ---------------------------------------------------------------------------
# Marker interpretation (shared by the table cells, color coding, and legend)
# ---------------------------------------------------------------------------

# Display glyphs
SYM_STAR = '★'          # red star: joint, teaching two seminars
SYM_HALF = '◗'          # half circle: joint appointment / co-teaching
SYM_DOUBLE = '✓✓'  # double check: multiple sections
SYM_CHECK = '✓'         # single check: taught that year


def marker_symbol(marker):
    """Return the table glyph for a raw SAVED marker, or '' for a blank cell."""
    value = str(marker or '').strip()
    if not value:
        return ''
    if 'XXSTAR' in value.upper():
        return SYM_STAR
    if '*' in value:
        return SYM_HALF
    if value == 'XX':
        return SYM_DOUBLE
    return SYM_CHECK


def marker_color_key(marker):
    """Return the color bucket for a marker, matching the cell text color.

    B (retired / no longer at Harvard) wins over the rank annotations.
    """
    value = str(marker or '').strip()
    if 'B' in value:
        return 'black'
    if '!!' in value:
        return 'pink'
    if '!' in value:
        return 'grey'
    return 'crimson'


def legend_key(marker):
    """Which legend line a marker belongs to, or None if it has no line.

    Mirrors marker_symbol() and marker_color_key() so the legend lists exactly
    the rules that actually appear in the table. Grey/pink half circles are an
    edge case with no dedicated legend entry (as before) and return None.
    """
    symbol = marker_symbol(marker)
    if not symbol:
        return None
    color = marker_color_key(marker)
    if symbol == SYM_STAR:
        return 'star'
    if symbol == SYM_DOUBLE:
        return 'double'
    if symbol == SYM_HALF:
        return {'black': 'joint_black', 'crimson': 'joint'}.get(color)
    return {'crimson': 'current', 'grey': 'entry',
            'pink': 'higher', 'black': 'check_black'}[color]


# ---------------------------------------------------------------------------
# PDF Builder
# ---------------------------------------------------------------------------

# Font sizes and padding decide how wide the table comes out. The width probe
# in create_pdf5 measures with exactly these, so a measurement can never drift
# from what is actually drawn.
HEADER_FONT_SIZE = 10
NAME_FONT_SIZE = 11
MARK_FONT_SIZE = 12
CELL_PADDING = 4


def create_pdf5(dept, output_path, headers, dept_data):
    """Create PDF 5 (Faculty Teaching Table) for a specific department.

    Generates a landscape-oriented PDF with:
    - A "Faculty Teaching Seminars by Name" main title
    - The department full name as a crimson subtitle
    - A table with one row per faculty member, showing:
        * Professor name (bold)
        * Rank (abbreviated if needed)
        * Checkmark grid for each academic year
    - Color-coded checkmarks (crimson, grey, pink) per teaching marker
    - A legend explaining the symbols and colors

    Faculty are sorted by rank hierarchy (University Professors first) and
    alphabetically within each rank group.

    Args:
        dept: Department acronym (e.g., 'HIST').
        output_path: Full file path for the output PDF.
        headers: List of column headers from SAVED.xlsx.
        dept_data: List of row dicts for this department from SAVED.xlsx.

    Returns:
        The oldest year columns left off because the table would not have fit on
        the page (see trim_years_to_fit), oldest first. Empty for the
        departments that show every year, which is nearly all of them.

    Table styling:
        - Crimson header row with white text
        - Alternating white/light-grey row backgrounds
        - Grey grid lines
        - Color-coded checkmarks for different teaching marker types
    """
    department_full_name = ACRONYM_TO_FULL.get(dept, dept)

    doc = SimpleDocTemplate(output_path, pagesize=landscape(A4),
                            rightMargin=0.5 * inch, leftMargin=0.5 * inch,
                            topMargin=0.5 * inch, bottomMargin=0.5 * inch)
    story = []
    styles = getSampleStyleSheet()

    # Department name is the primary header: large and crimson, on top.
    dept_style = ParagraphStyle('DeptStyle', parent=styles['Title'],
                                fontSize=24, spaceAfter=10, alignment=1,
                                textColor=colors.HexColor('#8B0000'))
    # Report name is the secondary line: smaller and black, below.
    subtitle_style = ParagraphStyle('SubtitleStyle', parent=styles['Heading2'],
                                    fontSize=20, spaceAfter=20, alignment=1,
                                    textColor=colors.black)

    story.append(Paragraph(department_full_name, dept_style))
    story.append(Paragraph("Faculty Teaching Seminars by Name", subtitle_style))

    # Identify year columns (everything except Professor, Rank, Department)
    year_columns = [h for h in headers
                    if str(h).strip().lower() not in ('id', 'professor', 'rank', 'department')]

    # Sort faculty by rank hierarchy, then alphabetically within each rank
    def sort_key(row):
        return (get_rank_sort_key(row.get('Rank', '')), str(row.get('Professor', '')).lower())

    sorted_data = sorted(dept_data, key=sort_key)

    # --- Build table data ---
    def build_table(years):
        """The whole table, header row first, for a given set of year columns."""
        rows = [['Professor', 'Rank'] + [str(y).strip() for y in years]]

        for row in sorted_data:
            table_row = []

            # Professor name
            professor = row.get('Professor', '')
            table_row.append(str(professor))

            # Rank (converted to full name, then abbreviated for display)
            rank = row.get('Rank', '')
            if rank:
                full_rank = get_rank_full_name(rank)
                abbreviated = get_rank_abbreviation(full_rank)
                clean = abbreviated.replace('<br/>', '\n').replace('<b>', '').replace('</b>', '')
                table_row.append(clean)
            else:
                table_row.append('')

            # Year columns: convert data markers to display symbols
            for year in years:
                table_row.append(marker_symbol(row.get(year, '')))

            rows.append(table_row)

        return rows

    # The style commands that decide the column widths. The probe below and the
    # real table both start from these, so what is measured is what is drawn.
    layout_style = [
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), HEADER_FONT_SIZE),
        ('FONTNAME', (0, 1), (1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (1, -1), NAME_FONT_SIZE),
        ('FONTNAME', (2, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (2, 1), (-1, -1), MARK_FONT_SIZE),
        ('LEFTPADDING', (0, 0), (-1, -1), CELL_PADDING),
        ('RIGHTPADDING', (0, 0), (-1, -1), CELL_PADDING),
    ]

    def fits_on_page(years):
        """Does the table built from these years stay inside the paper?

        An over-wide table is centred, so it only loses text once it reaches the
        page edge, not merely the margin. Measured with ReportLab's own layout
        rather than estimated, so this keeps working as years are added.
        """
        probe = Table(build_table(years), repeatRows=1)
        probe.setStyle(TableStyle(layout_style))
        return probe.wrap(doc.width, doc.height)[0] <= doc.pagesize[0]

    # Departments that fit keep every year. The ones that would be cut off start
    # a year later; from here on year_columns is only the years actually shown,
    # so the cell colouring and the legend follow the table.
    all_years = year_columns
    year_columns = trim_years_to_fit(all_years, fits_on_page)
    left_off = dropped_years(all_years, year_columns)

    table_data = build_table(year_columns)
    table = Table(table_data, repeatRows=1)

    # --- Style the table ---
    style = TableStyle(layout_style + [
        # Header row: crimson background, white bold text
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B0000')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),

        # Name and Rank columns: left-aligned
        ('ALIGN', (0, 1), (1, -1), 'LEFT'),

        # Year data columns: centered
        ('ALIGN', (2, 1), (-1, -1), 'CENTER'),

        # Grid and alternating row colors
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),

        # Cell padding and alignment
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])

    # --- Apply color coding to checkmarks based on original data markers ---
    # B suffix = black (Retired or No Longer at Harvard)
    # ! = grey (FYSP Entry Rank), !! = pink (FYSP Higher Rank)
    # Otherwise = crimson (current rank)
    marker_colors = {
        'black': colors.black,                    # B: retired / no longer at Harvard
        'pink': colors.HexColor('#FF6B6B'),       # !!: FYSP higher rank
        'grey': colors.grey,                      # !: FYSP entry rank
        'crimson': colors.HexColor('#8B0000'),    # standard (incl. the red star)
    }
    for row_idx in range(1, len(table_data)):
        for col_idx in range(2, len(table_data[row_idx])):
            cell = table_data[row_idx][col_idx]
            original = str(sorted_data[row_idx - 1].get(year_columns[col_idx - 2], '') or '').strip()

            if not cell:
                continue

            style.add('TEXTCOLOR', (col_idx, row_idx), (col_idx, row_idx),
                      marker_colors[marker_color_key(original)])

    table.setStyle(style)
    story.append(table)

    # --- Legend ---
    story.append(Spacer(1, 20))

    legend_title_style = ParagraphStyle('LegendTitle', parent=styles['Heading3'],
                                        fontSize=16, spaceAfter=10, alignment=0,
                                        textColor=colors.HexColor('#8B0000'))
    legend_style_red = ParagraphStyle('LegendRed', parent=styles['Normal'],
                                      fontSize=13, spaceAfter=5, leftIndent=20,
                                      textColor=colors.HexColor('#8B0000'))
    legend_style_grey = ParagraphStyle('LegendGrey', parent=styles['Normal'],
                                       fontSize=13, spaceAfter=5, leftIndent=20,
                                       textColor=colors.grey)
    legend_style_pink = ParagraphStyle('LegendPink', parent=styles['Normal'],
                                       fontSize=13, spaceAfter=5, leftIndent=20,
                                       textColor=colors.HexColor('#FF6B6B'))

    legend_style_black = ParagraphStyle('LegendBlack', parent=styles['Normal'],
                                        fontSize=13, spaceAfter=5, leftIndent=20,
                                        textColor=colors.black)

    # Only show the rules whose markers actually appear in this department's
    # table, in a fixed canonical order.
    legend_entries = [
        ('current',     "\u2713 = Current Rank Years Taught", legend_style_red),
        ('entry',       "\u2713 = FYSP Entry Rank", legend_style_grey),
        ('higher',      "\u2713 = FYSP Higher Rank", legend_style_pink),
        ('double',      "\u2713\u2713 = Multiple Sections Taught (2.0 score)", legend_style_red),
        ('star',        "\u2605 = Joint, Teaching Two Seminars", legend_style_red),
        ('joint',       "\u25d7 = Joint Department or Co-teaching (0.5 score)", legend_style_red),
        ('joint_black', "\u25d7 = Retired or No Longer at Harvard", legend_style_black),
        ('check_black', "\u2713 = Retired or No Longer at Harvard", legend_style_black),
    ]

    present_keys = set()
    for row in sorted_data:
        for year in year_columns:
            key = legend_key(row.get(year, ''))
            if key:
                present_keys.add(key)

    if present_keys:
        story.append(Paragraph("Legend:", legend_title_style))
        for key, text, sty in legend_entries:
            if key in present_keys:
                story.append(Paragraph(text, sty))

    doc.build(story)
    return left_off
