"""PDF 4: Seminars Taught per Year per Rank - Aggregated table."""

import os
import openpyxl
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch

from .shared import ACRONYM_TO_FULL, trim_years_to_fit, dropped_years

RANK_CATEGORIES = [
    "Professor & University Professor",
    "Associate Professor",
    "Assistant Professor",
    "SL/SP/PoP/PiR",
    "Emeritus & Emerita",
    "Lecturer",
    "Visiting (all levels)",
]

RANK_ACRONYM_EXPLANATIONS = {
    "SL/SP/PoP/PiR": "Senior Lecturer/Senior Preceptor/Professor of the Practice/Professor in Residence"
}


def categorize_rank(rank):
    """Categorize a rank into one of the predefined categories."""
    if not rank:
        return "Visiting (all levels)"

    original_rank_lower = rank.lower()

    if '/' in rank:
        rank_parts = rank.split('/')
        rank = rank_parts[-1].strip()
        if not rank:
            return "Visiting (all levels)"

    rank_lower = rank.lower()

    if 'visiting' in original_rank_lower:
        return "Visiting (all levels)"
    if 'associate professor' in rank_lower:
        return "Associate Professor"
    if 'assistant professor' in rank_lower:
        return "Assistant Professor"
    if any(kw in rank_lower for kw in ['senior lecturer', 'senior preceptor',
                                        'professor of the practice', 'professor in residence',
                                        'prof. of the practice']):
        return "SL/SP/PoP/PiR"
    if any(kw in rank_lower for kw in ['emeritus', 'emerita']):
        return "Emeritus & Emerita"
    if any(kw in rank_lower for kw in ['lecturer', 'instructor']) and 'senior' not in rank_lower:
        return "Lecturer"
    if 'professor' in rank_lower or rank_lower == 'prof.':
        if not any(kw in rank_lower for kw in ['associate', 'assistant', 'visiting',
                                                'senior lecturer', 'senior preceptor',
                                                'professor of the practice',
                                                'professor in residence',
                                                'emeritus', 'emerita']):
            return "Professor & University Professor"

    return "Visiting (all levels)"


def read_excel_data_by_department(file_path):
    """Read the Excel file and return data grouped by department."""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value) if cell.value else f"Column_{cell.column}")

    data_by_dept = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        row_data = {}
        for i, cell in enumerate(row):
            if i < len(headers):
                row_data[headers[i]] = cell.value if cell.value is not None else ""

        dept = row_data.get('Department', '')
        if dept:
            data_by_dept.setdefault(dept, []).append(row_data)

    return headers, data_by_dept


def format_number(value):
    """Format numbers consistently: 1.0, 1.5, 2.0, etc."""
    if value == 0:
        return ''
    elif value == int(value):
        return f"{int(value)}.0"
    else:
        return f"{value:.1f}"


def year_columns_of(data):
    """The academic-year columns of the master sheet, oldest first."""
    return [k for k in data[0].keys()
            if str(k).strip().lower() not in ('id', 'department', 'professor', 'rank')]


def abbreviate_year(year):
    """Year heading as this table prints it: '2013-14' becomes '13-14'.

    A heading in any other shape is left exactly as it is.
    """
    if '-' not in year:
        return year
    start, _, end = year.partition('-')
    return f"{start[-2:] if len(start) == 4 else start}-{end[-2:] if len(end) == 4 else end}"


# "B" marks a professor who has retired, left Harvard, or died while teaching.
# It says something about the person today, not about whether the seminar ran,
# so it never changes a score: XB counts as X, and X*B or XB* as X*. Page 5
# still prints those markers in black, which is where the reader learns it.
def score_marker(value):
    """A marker with its status annotation removed, for scoring only."""
    text = str(value).strip()
    if text.upper() == 'XXSTAR':
        return text
    return text.replace('B', '').replace('b', '')


def aggregate_by_rank(data):
    """Aggregate teaching data by rank categories."""
    rank_data = {rank: {} for rank in RANK_CATEGORIES}

    year_columns = year_columns_of(data)
    abbreviated_year_columns = [abbreviate_year(y) for y in year_columns]

    for rank in RANK_CATEGORIES:
        for year in abbreviated_year_columns:
            rank_data[rank][year] = 0

    for row in data:
        rank = row.get('Rank', '')
        rank_text_clean = rank.replace('\n', ' ').replace('\r', ' ').strip() if isinstance(rank, str) else str(rank)
        rank_category = categorize_rank(rank)

        for i, year in enumerate(year_columns):
            value = score_marker(row.get(year, ''))
            abbr_year = abbreviated_year_columns[i]

            # Multi-rank handling
            if '/' in rank_text_clean and value in ('X!', 'X', 'X!!', 'X*!', 'X*'):
                rank_parts = [p.strip() for p in rank_text_clean.split('/')]
                if len(rank_parts) >= 2:
                    first_cat = categorize_rank(rank_parts[0])
                    second_cat = categorize_rank(rank_parts[1]) if len(rank_parts) > 1 else first_cat
                    last_cat = categorize_rank(rank_parts[-1])

                    if value == 'X!':
                        rank_data[first_cat][abbr_year] += 1
                        continue
                    elif value == 'X!!':
                        rank_data[second_cat][abbr_year] += 1
                        continue
                    elif value == 'X*!' or value == 'X*':
                        rank_data[first_cat][abbr_year] += 0.5
                        continue
                    elif value == 'X':
                        rank_data[last_cat][abbr_year] += 1
                        continue

            # Normal aggregation
            if str(value).strip().upper() == 'XXSTAR':
                # Joint appointment teaching two seminars: score 1 for each
                # department row it is joint of. A joint-of-two case adds 1 in
                # each department (2 total program-wide), matching a non-joint
                # two-seminar case ('XX', which scores 2 in its one department).
                rank_data[rank_category][abbr_year] += 1
            elif value == 'X':
                rank_data[rank_category][abbr_year] += 1
            elif value == 'XX':
                rank_data[rank_category][abbr_year] += 2
            elif value in ('X*', 'X**', 'X***', 'X* **'):
                rank_data[rank_category][abbr_year] += 0.5
            elif value in ('X!', 'X!!'):
                rank_data[rank_category][abbr_year] += 1
            elif value == 'X*!':
                rank_data[rank_category][abbr_year] += 0.5

    return rank_data, abbreviated_year_columns


# Font sizes and padding decide how wide the table comes out. The width probe
# in create_pdf4 measures with exactly these, so a measurement can never drift
# from what is actually drawn.
HEADER_FONT_SIZE = 10
RANK_FONT_SIZE = 12
VALUE_FONT_SIZE = 12
RANK_PADDING = 12
CELL_PADDING = 4


def layout_style(row_count):
    """The style commands that decide the column widths.

    The probe and the real table both start from these, so what is measured is
    what is drawn. The Total row is bold, and bold digits are wider, so its
    font belongs here too.
    """
    total_idx = row_count - 1
    return [
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), HEADER_FONT_SIZE),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (0, -1), RANK_FONT_SIZE),
        ('LEFTPADDING', (0, 1), (0, -1), RANK_PADDING),
        ('RIGHTPADDING', (0, 1), (0, -1), RANK_PADDING),
        ('FONTNAME', (1, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (1, 1), (-1, -1), VALUE_FONT_SIZE),
        ('LEFTPADDING', (1, 0), (-1, -1), CELL_PADDING),
        ('RIGHTPADDING', (1, 0), (-1, -1), CELL_PADDING),
        ('FONTNAME', (0, total_idx), (-1, total_idx), 'Helvetica-Bold'),
        ('FONTSIZE', (0, total_idx), (-1, total_idx), VALUE_FONT_SIZE),
    ]


def create_pdf4(dept, output_path, saved_data):
    """Create PDF 4 (Rank Aggregator) for a specific department.

    saved_data: list of row dicts for this department from FYSP Master

    Returns:
        The oldest year columns left off because the table would not have fit
        across the page, oldest first and named as the master spreadsheet names
        them ("2013-14"). Empty for the departments that show every year, which
        is nearly all of them.
    """
    department_full_name = ACRONYM_TO_FULL.get(dept, dept)

    rank_data, year_columns = aggregate_by_rank(saved_data)
    # Headings are abbreviated for the table ('13-14') but reported to whoever
    # runs this in the spreadsheet's own spelling ('2013-14').
    full_year_names = dict(zip(year_columns, year_columns_of(saved_data)))

    year_totals = {}
    for year in year_columns:
        year_totals[year] = sum(rank_data[rank][year] for rank in RANK_CATEGORIES)

    doc = SimpleDocTemplate(output_path, pagesize=landscape(A4),
                            leftMargin=0.8 * inch, rightMargin=0.8 * inch,
                            topMargin=0.5 * inch, bottomMargin=0.5 * inch)
    story = []
    styles = getSampleStyleSheet()

    # Department name on top: larger, crimson.
    dept_style = ParagraphStyle('DeptTitle', parent=styles['Heading1'],
                                fontSize=20, spaceAfter=10, alignment=1,
                                textColor=colors.HexColor('#8B0000'))
    # Report name below: smaller, black.
    subtitle_style = ParagraphStyle('SubTitle', parent=styles['Heading2'],
                                    fontSize=16, spaceAfter=20, alignment=1)

    story.append(Paragraph(department_full_name, dept_style))
    story.append(Paragraph("Seminars Taught per Year by Rank", subtitle_style))

    # Build table
    def build_table(years):
        """The whole table, header row first, for a given set of year columns."""
        rows = [['Rank'] + list(years)]
        for rank in RANK_CATEGORIES:
            rows.append([rank.replace('/', ' ')]
                        + [format_number(rank_data[rank][y]) for y in years])
        rows.append(['Total'] + [format_number(year_totals[y]) for y in years])
        return rows

    def fits_on_page(years):
        """Does the table built from these years stay inside the paper?

        An over-wide table is centred, so it only loses text once it reaches
        the page edge, not merely the margin. Measured with ReportLab's own
        layout rather than estimated, so this keeps working as years are added.
        """
        rows = build_table(years)
        probe = Table(rows, repeatRows=1)
        probe.setStyle(TableStyle(layout_style(len(rows))))
        return probe.wrap(doc.width, doc.height)[0] <= doc.pagesize[0]

    # Departments that fit keep every year. The ones that would be cut off at
    # both edges start a year later instead.
    all_years = year_columns
    year_columns = trim_years_to_fit(all_years, fits_on_page)
    left_off = [full_year_names.get(y, y)
                for y in dropped_years(all_years, year_columns)]

    table_data = build_table(year_columns)
    table = Table(table_data, repeatRows=1)

    style = TableStyle(layout_style(len(table_data)) + [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B0000')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#f8f9fa')),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ])

    total_idx = len(table_data) - 1
    style.add('BACKGROUND', (0, total_idx), (-1, total_idx), colors.HexColor('#e6e6e6'))

    table.setStyle(style)
    story.append(table)

    # Legend
    story.append(Spacer(1, 20))
    legend_title_style = ParagraphStyle('LegendTitle', parent=styles['Heading3'],
                                        fontSize=14, spaceAfter=8, alignment=0)
    legend_style = ParagraphStyle('Legend', parent=styles['Normal'],
                                  fontSize=13, spaceAfter=5, leftIndent=20)

    story.append(Paragraph("Legend", legend_title_style))
    story.append(Paragraph("\u2022 Joint Instruction or Co-Teaching = 0.5", legend_style))
    for acronym, full_name in RANK_ACRONYM_EXPLANATIONS.items():
        story.append(Paragraph(f"\u2022 {acronym} = {full_name}", legend_style))

    doc.build(story)
    return left_off
